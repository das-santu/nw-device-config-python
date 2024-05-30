import paramiko
from openpyxl import load_workbook
from openpyxl.styles import Font, colors

# Input file name with location
input_xl_file = 'files/devices.xlsx'

# Load the Excel file
wb = load_workbook(input_xl_file)
device_sheet = wb['devices']
result_sheet = wb['result']

result_row = 1
# Iterate over the rows in the Excel file
for row in device_sheet.iter_rows(min_row=2, values_only=True):
    ip, username, password, config = row
    result_row += 1
    
    if ip:
        # Printing configuring device
        print(f'Configuring Device: {ip}')
        
        try:
            # Connect to the device
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(ip, username=username, password=password, timeout=20)

            # Send the configuration commands
            stdin, stdout, stderr = ssh.exec_command(config)

            # Check if the configuration was successful
            if stderr.read():
                result = 'Not Successful'
                font = Font(color="EA2F15")
            else:
                result = 'Successful'
                font = Font(color="00B334")
                
        except paramiko.AuthenticationException:
            result = 'Authentication Failed'
            font = Font(color="EA2F15")
        except paramiko.SSHException:
            result = 'SSH Connection Failed'
            font = Font(color="EA2F15")
        except Exception:
            result = 'Unreachable/Offline'
            font = Font(color="EA2F15")

        # Write the result to the Excel file
        result_sheet.cell(row=result_row, column=1, value=ip)
        result_sheet.cell(row=result_row, column=2, value=result).font = font

        # Close the SSH connection
        ssh.close()

# Save the Excel file
wb.save(input_xl_file)
